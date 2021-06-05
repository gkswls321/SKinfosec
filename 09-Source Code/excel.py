from datetime import date
import datetime
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border
from bum import *
import pandas as pd
# -- 엑셀 만들기 --

wb = openpyxl.Workbook() # 새 워크북(임시 엑셀 파일) 생성
sheet1 = wb.create_sheet("sheet_1", 0) # sheet_1가 맨 앞에 생성
sheet2 = wb.create_sheet("sheet_2", 1) # sheet_2가 2번째 시트에 생성
sheet3 = wb.create_sheet("sheet_3", 2) # sheet_3가 3번째 시트에 생성
sheet2.title = 'change_sheet2'
sheet3.title = 'CT_LOG'
sheet1['A1'] = 123  
sheet1['A2'] = 456

sheet3['B5'] = 'CT_LOG'
print(type(len(x)))
for i in range(len(x)):
    sheet3.cell(row = 6, column = 2 + i, value = x[i])
    sheet3.cell(row = 7, column = 2 + i, value=num[i])



# sheet3['B3'] = raw_data


sheet1.append(['first','second']) # 마지막 행에 이어서 추가

sheet2.cell(1,1).value = 'second sheet'

ft = Font(name='맑은고딕', color='ffe666', size=25,bold=True, italic=True, strikethrough = True, underline='single')
sheet1['A1'].font = ft

sheet1.merge_cells('B1:D1')
# ws.merge_cells(2,1,2,4) # 2행 1열셀 ~ 2행 4열셀까지 병합
# ws.unmerge_cells # 병합 해제

sheet1['B1'] = 'merge cells' # 병합된 셀에 값을 넣을 땐 첫번째 셀을 기준으로
sheet1['B1'].font = Font(name='맑은고딕',size=20)
sheet1['B1'].alignment = Alignment(horizontal='center', vertical='bottom') # 가운데 정렬

# 테두리 설정
# sheet1['B1'].border = Border(left=borders.Side(style='thick')) # 이거 안되는데?? 이유는 모름

# 이미지 삽입
img = openpyxl.drawing.image.Image('./main.png')
img.anchor = 'B1'
sheet2.add_image(img)


img = openpyxl.drawing.image.Image('./2021-06-03-18-32-00_CT_LOG.jpg')
img.anchor = 'B10'
sheet3.add_image(img)
# 날짜
today = date.today()

now = datetime.datetime.now()
nowDate = str(now.strftime('%Y-%m-%d'))
nowTime = str(now.strftime('%H-%M-%S'))

today_y = today.year
today_m = today.month
today_d = today.day

# globals()['{0}_{1}_excel_report.xlsx'.format(nowDate,nowTime)]

filepath = './{0}-{1}_excel_report.xlsx'.format(nowDate,nowTime)
wb.save(filepath) # 

# wb.save(filename="c\test\test01.xlsx")
# C:/Users/dolif/Desktop/프로젝트

print('------','complete','----------',sep='\n')
# lwb = oxl.load_workbook("exist.xls") # 기존에 있는 엑셀 파일 열기

