from datetime import date
import datetime
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side
import sys
sys.path.append('C:/Users/dolif/Desktop/프로젝트/')

from apache2 import ap_x_values
from apache2 import ap_values ### apache2에서 원하는 ap_values 값 가져옴 ######

# print(APACHE_T_2)

# -- DATE ------------
today = date.today()

now = datetime.datetime.now()
nowDate = str(now.strftime('%Y-%m-%d'))
nowTime = str(now.strftime('%H'))


# -- make excel --
wb = openpyxl.Workbook() # 새 워크북(임시 엑셀 파일) 생성
sheet1 = wb.create_sheet("EC2_REPORT", 0) # sheet_1가 맨 앞에 생성
sheet2 = wb.create_sheet("APACHE_WEB_REPORT", 1) # sheet_2가 2번째 시트에 생성
sheet3 = wb.create_sheet("MYSQL_REPORT", 2)
sheet4 = wb.create_sheet("CLOUD TRAIL_REPORT", 3)

# insert graph
def insert_image(a,img,b,c,d):
    img.anchor = d

    unit_cm= b
    unit_cm2 = c

    unit_inch= round((unit_cm/2.54)*10,0)#  change to float
    unit_inch= round((unit_cm/2.54)*100)  #  change to integer

    unit_inch2= round((unit_cm2/2.54)*10,0)#  change to float
    unit_inch2= round((unit_cm2/2.54)*100)  #  change to integer

    img.height = unit_inch
    img.width = unit_inch2
    a.add_image(img)


# make title ---
def make_title(a):
    SHEET_border = Border(left=Side(border_style='thick', color='000000'), 
            right=Side(border_style='thick', color='000000'),
			top=Side(border_style='thick', color='000000'),
			bottom=Side(border_style='thick', color='000000')) 
    
    a.merge_cells('A1:H1')
    a['A1'] = 'EVENTSHOP REPORT'
    a['A1'].font = Font(size=40,bold=True)
    a['A2'] = 'Today : ' 
    a['b2'] = nowDate

    insert_image(a,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/adt.png'),2,4,'g2')
    # a['A1:G1'].border = SHEET_border

make_title(sheet1)  
make_title(sheet2)
make_title(sheet3)
make_title(sheet4)


# insert_image(sheet1,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_apache_web_graph.png'.format(nowDate,nowTime)))

insert_image(sheet2,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_apache_web_graph.png'.format(nowDate,nowTime)),12,16,'a6')

# insert_image(sheet3,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_apache_web_graph.png'.format(nowDate,nowTime)),12,16,'a6')

# insert_image(sheet4,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_apache_web_graph.png'.format(nowDate,nowTime)),12,16,'a6')

# sheet1.append(['first','second']) # A1 -> A2 마지막 행에 이어서 추가


#-------- A30부터 원하는 내용추가하기 # LIST 내용 넣기
ap_x_values # x축
ap_values # y축

for i in range(len(ap_x_values)):
    sheet2.cell(30, column = 1 + i, value = ap_x_values[i])
    sheet2.cell(31, column = 1 + i, value= ap_values[i])

ft = Font(name='Arial', size=25,bold=True, italic=True, strikethrough = True, underline='single')
# sheet1['A2'] = 'HI'
# sheet1['A2'].font = ft

# sheet1['C1'] = 'merge cells' # 병합된 셀에 값을 넣을 땐 첫번째 셀을 기준으로
# sheet1['C1'].font = Font(name='Arial',size=20)
# sheet1['C1'].alignment = Alignment(horizontal='center', vertical='bottom') # 가운데 정렬

# ---------- save ----------
filepath = 'C:/Users/dolif/Desktop/프로젝트/{0}-{1}_excel_report.xlsx'.format(nowDate,nowTime)
wb.save(filepath) # 

print('------','complete','----------',sep='\n')


