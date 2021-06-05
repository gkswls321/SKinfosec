from datetime import date
import datetime
from typing import NewType
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side

# import sys
# sys.path.append('C:/Users/dolif/Desktop/프로젝트/excel') 
from apache2 import ap_x_values, ap_values
from mysql_log import mysql_Warning
from WEB1_net_log import net_time_data2
from WEB2_net_log import net2_time_data2
from WEB1_cpu_net import cpu_time_data2
from WEB2_cpu_net import cpu2_time_data2

# -- DATE ------------
today = date.today()

now = datetime.datetime.now()
nowDate = str(now.strftime('%Y-%m-%d'))
nowTime = str(now.strftime('%H'))


# -- make excel --
wb = openpyxl.Workbook() # create workbook
sheet1 = wb.create_sheet("WEB_REPORT", 0) # First sheet1 create 
sheet2 = wb.create_sheet("APACHE_WEB_ERROR_REPORT", 1) # Second sheet2 create 
sheet3 = wb.create_sheet("MYSQL_ERROR_REPORT", 2)


# insert graph
def insert_image(a,img,b,c,d):
    img.anchor = d

    unit_cm= b
    unit_cm2 = c

    unit_inch= round((unit_cm/2.54)*10,0)#  change to float
    unit_inch= round((unit_cm/2.54)*100)  #  change to integer

    unit_inch2= round((unit_cm2/2.54)*10,0)#  change to float
    unit_inch2= round((unit_cm2/2.54)*100)  #  change to integer

    img.height = unit_inch
    img.width = unit_inch2
    a.add_image(img)

# make title ---
def make_title(a):
    SHEET_border = Border(left=Side(border_style='thick', color='000000'), 
            right=Side(border_style='thick', color='000000'),
			top=Side(border_style='thick', color='000000'),
			bottom=Side(border_style='thick', color='000000')) 
    
    a.merge_cells('A1:H1')
    a['A1'] = 'EVENTSHOP REPORT'
    a['A1'].font = Font(size=40,bold=True)
    a['A2'] = 'Today : ' 
    a['b2'] = nowDate

make_title(sheet1)  
make_title(sheet2)
make_title(sheet3)

insert_image(sheet1,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_WEB1_CPU_MAX_graph.png'.format(nowDate,nowTime)),9,15,'A6')

insert_image(sheet1,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_WEB2_CPU_MAX_graph.png'.format(nowDate,nowTime)),9,15,'A22')

insert_image(sheet1,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_WEB1_NET_MAX_graph.png'.format(nowDate,nowTime)),9,15,'J6')

insert_image(sheet1,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_WEB2_NET_MAX_graph.png'.format(nowDate,nowTime)),9,15,'J22')

insert_image(sheet2,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_apache_web_graph.png'.format(nowDate,nowTime)),12,16,'A6')

insert_image(sheet3,openpyxl.drawing.image.Image('C:/Users/dolif/Desktop/프로젝트/{0}-{1}_MYSQL_ERROR_TYPE_graph.png'.format(nowDate,nowTime)),14,20,'A6')

#-------- INSERT WANT DATE TO A30 
# INSERT LIST DATA!
ap_x_values # x
ap_values # y

for i in range(len(ap_x_values)):
    sheet2.cell(30, column = 1 + i, value = ap_x_values[i])
    sheet2.cell(31, column = 1 + i, value= ap_values[i])

# for i in range(len(mysql_Warning)):
#     sheet3.cell(36+i , column = 1, value= mysql_Warning[i])

# ---------- save ----------
filepath = 'C:/Users/dolif/Desktop/프로젝트/{0}-{1}_excel_report.xlsx'.format(nowDate,nowTime)
wb.save(filepath) # 

print('------','complete','----------',sep='\n')
